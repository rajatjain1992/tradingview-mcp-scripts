#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Download the last N months of EQUITY-scheme monthly portfolio disclosures for
SBI, ICICI Prudential and HDFC mutual funds, parse them into one normalized
table, save to a local SQLite database, and load into BigQuery.

Source pages (AMFI -> per-AMC monthly portfolio disclosure):
    SBI   : https://www.sbimf.com/portfolios
            One combined "All Schemes Monthly Portfolio" XLSX per month
            (multi-sheet workbook, one sheet per scheme, with an "Index"
            sheet mapping scheme short-code -> full scheme name).
            Direct URL pattern (confirmed, no auth):
            https://www.sbimf.com/docs/default-source/scheme-portfolios/
                all-schemes-monthly-portfolio---as-on-{DDth}-{month}-{YYYY}.xlsx

    ICICI : https://www.icicipruamc.com/media-center/downloads
            ?currentTabFilter=Disclosures&subCatTabFilter=MonthlyPortfolioDisclosures
            The "Download" buttons are JS-triggered and resolve to either a
            single combined multi-sheet workbook OR a ZIP of one-xlsx-per-
            scheme (observed both forms across months in practice) -- there
            is no stable, guessable static URL, so this script does NOT
            re-download ICICI automatically. It expects the per-scheme xlsx
            files to already be present, one subfolder per month, under
            cache/mf_portfolios/icici_zip_{apr,may,jun} (or matching
            --icici-dir naming) -- obtained by visiting the page above with
            a browser, clicking "Download" for each month, and unzipping.

    HDFC  : https://www.hdfcfund.com/statutory-disclosure/portfolio/monthly-portfolio
            One XLSX per individual scheme, hosted directly (no auth) at:
            https://files.hdfcfund.com/s3fs-public/{upload-yyyy-mm}/
                Monthly%20{Scheme%20Name}%20-%20{DD}%20{Month}%20{YYYY}.xlsx
            The upload-month folder is (in practice) the calendar month
            AFTER the portfolio month (e.g. the "30 June 2026" portfolio is
            uploaded under .../2026-07/) -- this script tries that offset
            first, then a couple of neighbouring months as a fallback.

Equity-scheme ground truth used for filtering (see EQUITY_SCHEMES below):
    - SBI: curated from the workbook's own "Index" sheet (scheme code ->
      full scheme name), keeping actively-managed equity categories
      (large/mid/small/flexi/multi/focused/value/contra/dividend
      yield/ELSS/sectoral-thematic) and excluding debt/liquid/gilt/
      overnight/hybrid/arbitrage/multi-asset/FoF/ETF/index schemes.
    - ICICI: curated from icicipruamc.com's own "Equity Funds" explorer
      page (https://www.icicipruamc.com/mutual-fund/equity-funds), which
      lists every actively-managed equity scheme ICICI itself classifies
      as "Equity Funds".
    - HDFC: curated from hdfcfund.com's own "Equity" filtered fund explorer
      (https://www.hdfcfund.com/explore/mutual-funds/equity, 21 schemes)
      PLUS "HDFC ELSS Tax Saver Fund" added by hand -- HDFC's own explorer
      puts ELSS in a separate "Tax Saver" filter bucket outside the
      default "Equity" category view, but SEBI's own regulatory scheme
      classification treats ELSS as an equity scheme category. This is a
      judgment call -- flagged in the script's final report so it can be
      overridden (pass --no-elss to exclude HDFC ELSS Tax Saver Fund).

Usage:
    # download SBI + HDFC directly, expect ICICI already cached, parse
    # everything, write sqlite, load to BigQuery
    python mf_equity_portfolios_to_bigquery.py --months 3 --to-bigquery

    # just parse what's already cached/downloaded, write sqlite only
    python mf_equity_portfolios_to_bigquery.py --months 3 --out mf_equity_portfolios.db
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import requests

PROJECT = "rajat-trade"
DATASET = "mutual_fund_data"
TABLE = "equity_portfolio_disclosures"
TABLE_ID = f"{PROJECT}.{DATASET}.{TABLE}"

SCRIPT_DIR = Path(__file__).parent
CACHE_DIR = SCRIPT_DIR / "cache" / "mf_portfolios"
DEFAULT_DB = SCRIPT_DIR / "mf_equity_portfolios.db"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
}

# ---------------------------------------------------------------------------
# Month targets: today is 2026-07-29 -> last 3 published monthly disclosures
# are as-on Apr 30, May 31, Jun 30 2026 (July's isn't published yet).
# ---------------------------------------------------------------------------
MONTHS = [
    {"date": "2026-04-30", "day": "30th", "day_num": "30", "month_name": "april", "month_title": "April", "sbi_sfvrsn": "fbc4d18a_2"},
    {"date": "2026-05-31", "day": "31st", "day_num": "31", "month_name": "may", "month_title": "May", "sbi_sfvrsn": "1e792ce4_2"},
    {"date": "2026-06-30", "day": "30th", "day_num": "30", "month_name": "june", "month_title": "June", "sbi_sfvrsn": "2891d9a4_2"},
]

# ---------------------------------------------------------------------------
# SBI: scheme short-code -> equity scheme, curated from the workbook's own
# "Index" sheet (scheme code -> full name), keeping actively managed equity
# categories, excluding debt/liquid/gilt/hybrid/arbitrage/multi-asset/FoF/
# ETF/index schemes.
# ---------------------------------------------------------------------------
SBI_EQUITY_SHEET_CODES = {
    "SMEEF", "SLMF", "SLTEF", "SMGLF", "SCOF", "STOF", "SHOF", "SCF", "SFEF",
    "SMIDCAP", "SMCOMMA", "SFLEXI", "SBLUECHIP", "SIF", "SPSU", "SSCF",
    "SBFS", "SLTAF-IV", "SLTAF-V", "SLTAF-VI", "SEMVF", "SMCF", "SDYF",
    "SEOF", "SBI-AOF", "SIOF", "SQF", "SQLF", "SBIRIOS",
}

# ---------------------------------------------------------------------------
# ICICI Prudential: equity scheme names, curated from icicipruamc.com's own
# "Equity Funds" explorer (https://www.icicipruamc.com/mutual-fund/equity-funds).
# Matched fuzzily (case-insensitive substring) against per-scheme filenames.
# ---------------------------------------------------------------------------
ICICI_EQUITY_SCHEMES = [
    "Manufacturing Fund", "Business Cycle Fund", "Flexicap Fund",
    "Active Momentum Fund", "Banking & Financial Services Fund",
    "Bharat Consumption Fund", "Commodities Fund", "Conglomerate Fund",
    "Dividend Yield Equity Fund", "ELSS Tax Saver Fund",
    "Energy Opportunities Fund", "Equity Minimum Variance Fund",
    "ESG Exclusionary Strategy Fund", "Exports and Services Fund",
    "FMCG Fund", "Focused Equity Fund", "Housing Opportunities Fund",
    "India Opportunities Fund", "Infrastructure Fund", "Innovation Fund",
    "Large & Mid Cap Fund", "Large Cap Fund",
    "Long Term Wealth Enhancement Fund", "Midcap Fund", "MNC Fund",
    "Multicap Fund", "Pharma Healthcare and Diagnostics (P.H.D) Fund",
    "PSU Equity Fund", "Quality Fund", "Quant Fund",
    "Rural Opportunities Fund", "Smallcap Fund", "Technology Fund",
    "Transportation and Logistics Fund", "US Bluechip Equity Fund",
    "Value Fund",
]

# ---------------------------------------------------------------------------
# HDFC: equity scheme filename fragments, curated from hdfcfund.com's own
# "Equity" filtered fund explorer (21 schemes) + ELSS Tax Saver added by hand
# (SEBI classifies ELSS as an equity scheme category even though HDFC's own
# site groups it under a separate "Tax Saver" filter -- judgment call, can
# be excluded with --no-elss).
# ---------------------------------------------------------------------------
HDFC_EQUITY_SCHEMES = [
    "Value Fund", "Transportation and Logistics Fund", "Technology Fund",
    "Small Cap Fund", "Pharma and Healthcare Fund", "Multi Cap Fund",
    "MNC Fund", "Mid Cap Fund", "Manufacturing Fund", "Large Cap Fund",
    "Large and Mid Cap Fund", "Innovation Fund", "Infrastructure Fund",
    "Housing Opportunities Fund", "Focused Fund", "Flexi Cap Fund",
    "Dividend Yield Fund", "Defence Fund", "Consumption Fund",
    "Business Cycle Fund", "Banking  Financial Services Fund",
]
HDFC_ELSS_SCHEME = "ELSS Tax saver"

BLOCK_EXACT = {
    "total", "grand total", "sub total", "subtotal", "derivatives total",
    "grand total (aum)", "total net assets",
}
# Rows matching these mark the true end of the holdings table (as opposed to
# a "Total"/"Sub Total" subsection subtotal with more holdings still to come
# below it) -- seeing one resets the column map so narrative/NAV-table rows
# further down the sheet aren't mistakenly parsed as holdings.
TABLE_END_MARKERS = ["grand total", "total net assets"]

# Narrative/administrative text that shows up after (or between) holdings
# tables in the SEBI-mandated monthly disclosure format (NAV-per-unit
# tables, notes, dividend/bonus declarations, hedging/derivative
# disclosures, risk-o-meter, etc) -- seeing any of these resets the column
# map so nothing past this point is mistaken for a holding.
NARRATIVE_MARKERS = [
    "notes", "plan name", "nav per unit", "dividend declared", "bonus declared",
    "nav at the beginning", "nav as on", "details of security in default",
    "details of stock future", "hedging position", "illiquid equity shares",
    "non traded/unlisted", "investment in foreign securities", "outstanding exposure in derivative",
    "portfolio turnover", "average maturity", "repo transaction",
    "risk-o-meter", "riskometer", "total value and percentage",
]
BLOCK_SUBSTR = [
    "equity & equity related", "equity and equity related",
    "listed / awaiting listing", "listed/awaiting listing",
    "debt instrument", "money market instrument", "government securit",
    "securitised debt", "mutual fund units", "certificate of deposit",
    "commercial paper", "treasury bill",
]


# ===========================================================================
# Download helpers
# ===========================================================================

def download(url: str, dest: Path, session: requests.Session | None = None) -> bool:
    if dest.exists() and dest.stat().st_size > 0:
        return True
    dest.parent.mkdir(parents=True, exist_ok=True)
    sess = session or requests
    resp = sess.get(url, headers=HEADERS, timeout=60)
    if resp.status_code != 200:
        print(f"  FAILED ({resp.status_code}): {url}")
        return False
    dest.write_bytes(resp.content)
    return True


def download_sbi(month: dict) -> Path | None:
    fname = f"sbi_{month['date']}.xlsx"
    dest = CACHE_DIR / fname
    if dest.exists():
        return dest
    url = (
        "https://www.sbimf.com/docs/default-source/scheme-portfolios/"
        f"all-schemes-monthly-portfolio---as-on-{month['day']}-{month['month_name']}-2026.xlsx"
        f"?sfvrsn={month['sbi_sfvrsn']}"
    )
    print(f"  SBI {month['date']}: downloading...")
    ok = download(url, dest)
    return dest if ok else None


def download_hdfc(month: dict) -> Path:
    """Downloads each HDFC equity scheme's xlsx directly (no auth needed).
    Tries the upload-month-folder = portfolio-month+1 convention observed
    in practice, falling back to +0/+2 if that 404s."""
    out_dir = CACHE_DIR / "hdfc"
    out_dir.mkdir(parents=True, exist_ok=True)
    schemes = list(HDFC_EQUITY_SCHEMES) + [HDFC_ELSS_SCHEME]
    port_dt = datetime.strptime(month["date"], "%Y-%m-%d")
    day_label = f"{port_dt.day} {month['month_title']} {port_dt.year}"

    def month_offset(dt: datetime, n: int) -> str:
        y, m = dt.year, dt.month + n
        while m > 12:
            m -= 12
            y += 1
        while m < 1:
            m += 12
            y -= 1
        return f"{y:04d}-{m:02d}"

    with requests.Session() as session:
        for scheme in schemes:
            fname = f"Monthly HDFC {scheme} - {day_label}.xlsx"
            dest = out_dir / fname
            if dest.exists():
                continue
            from urllib.parse import quote
            ok = False
            for offset in (1, 0, 2):
                folder = month_offset(port_dt, offset)
                url = f"https://files.hdfcfund.com/s3fs-public/{folder}/{quote(fname)}"
                if download(url, dest, session):
                    ok = True
                    break
            if not ok:
                print(f"  HDFC {month['date']}: FAILED to find {scheme}")
    return out_dir


def find_icici_dir(month: dict) -> Path | None:
    """ICICI's download buttons are JS-triggered with no stable static URL
    (a browser session obtained either a single combined workbook or a ZIP
    of per-scheme files depending on month) -- this script expects the
    already-unzipped per-scheme xlsx files to be present under one of a few
    conventional cache subfolder names."""
    candidates = [
        CACHE_DIR / f"icici_zip_{month['month_name'][:3]}",
        CACHE_DIR / f"icici_{month['month_name'][:3]}",
        CACHE_DIR / f"icici_{month['date']}",
    ]
    for c in candidates:
        if c.is_dir() and any(c.glob("*.xlsx")):
            return c
    return None


# ===========================================================================
# Generic SEBI-format holdings table parser
# ===========================================================================

def _cell_text(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _is_blocked(name_lower: str) -> bool:
    if name_lower in BLOCK_EXACT:
        return True
    return any(b in name_lower for b in BLOCK_SUBSTR)


def parse_holdings_sheet(ws, amc: str, scheme_name: str, portfolio_date: str) -> list[dict]:
    """Generic parser for the SEBI-mandated monthly portfolio holdings table.
    Detects header rows by locating a cell containing "instrument" (the one
    reliable anchor across SBI/ICICI/HDFC's differently-ordered column
    layouts), builds a column map from keyword matches in that header row,
    then emits a holding for every following row that has a name and at
    least one of ISIN / quantity / market value populated, skipping known
    section-header and total/subtotal rows."""
    rows_out = []
    col_map = None
    name_col = None

    for row in ws.iter_rows(values_only=True):
        texts = [_cell_text(c) for c in row]

        # Detect a new header row: needs the "instrument" anchor cell PLUS at
        # least one other recognizable column (isin/quantity/market value/pct)
        # in the same row, so narrative sentences that merely mention the
        # word "instrument" aren't mistaken for a real header.
        header_idx = None
        for i, t in enumerate(texts):
            if "instrument" in t:
                header_idx = i
                break
        if header_idx is not None:
            candidate = {"isin": None, "industry": None, "quantity": None, "market_value": None, "pct": None}
            for i, t in enumerate(texts):
                if i == header_idx:
                    continue
                if "isin" in t:
                    candidate["isin"] = i
                elif "industry" in t or "rating" in t:
                    candidate["industry"] = i
                elif "quantity" in t:
                    candidate["quantity"] = i
                elif "market" in t and "value" in t:
                    candidate["market_value"] = i
                elif ("% to" in t) or ("%" in t and "nav" in t) or ("%" in t and "aum" in t):
                    candidate["pct"] = i
            if sum(1 for v in candidate.values() if v is not None) >= 2:
                name_col = header_idx
                col_map = candidate
            continue

        # Narrative/administrative lines (NAV tables, notes, disclosures) mark
        # that we've left the holdings zone -- reset so nothing further in
        # this sub-section gets mistaken for a holding under a stale header.
        if any(any(marker in t for marker in NARRATIVE_MARKERS) for t in texts):
            col_map = None
            continue

        if col_map is None:
            continue

        name = row[name_col] if name_col < len(row) else None
        if name is None or not str(name).strip():
            continue
        name_str = str(name).strip()
        name_lower = name_str.lower()
        if _is_blocked(name_lower):
            if any(marker in name_lower for marker in TABLE_END_MARKERS):
                col_map = None  # stop parsing further rows in this sheet as holdings
            continue

        def get(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        isin_val = get("isin")
        isin_val = str(isin_val).strip() if isin_val not in (None, "") else None
        qty_val = _to_number(get("quantity"))
        mv_val = _to_number(get("market_value"))
        pct_val = _to_number(get("pct"))
        industry_val = get("industry")
        industry_val = str(industry_val).strip() if industry_val not in (None, "") else None

        if isin_val is None and qty_val is None and mv_val is None:
            continue  # section header / narrative row with no real holding data

        rows_out.append({
            "amc": amc,
            "scheme_name": scheme_name,
            "portfolio_date": portfolio_date,
            "isin": isin_val,
            "instrument_name": name_str,
            "industry_or_rating": industry_val,
            "quantity": qty_val,
            "market_value_lakhs": mv_val,
            "pct_to_nav": pct_val,
        })

    return rows_out


# ===========================================================================
# Per-AMC parsing
# ===========================================================================

def parse_sbi(path: Path, portfolio_date: str) -> list[dict]:
    if path is None or not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Index" not in wb.sheetnames:
        print(f"  SBI {portfolio_date}: no Index sheet, skipping")
        return []
    idx_ws = wb["Index"]
    code_to_name = {}
    for row in idx_ws.iter_rows(values_only=True):
        if row and len(row) >= 3 and row[1] and row[2]:
            code_to_name[str(row[1]).strip()] = str(row[2]).strip()

    all_rows = []
    for code in SBI_EQUITY_SHEET_CODES:
        if code not in wb.sheetnames:
            continue
        scheme_name = code_to_name.get(code, code)
        rows = parse_holdings_sheet(wb[code], "SBI", scheme_name, portfolio_date)
        all_rows.extend(rows)
    print(f"  SBI {portfolio_date}: {len(all_rows)} holding rows across "
          f"{len([c for c in SBI_EQUITY_SHEET_CODES if c in wb.sheetnames])} equity schemes")
    return all_rows


def _fuzzy_match_scheme(fname_stem: str, target: str) -> bool:
    a = re.sub(r"[^a-z0-9]", "", fname_stem.lower())
    b = re.sub(r"[^a-z0-9]", "", target.lower())
    return b in a


def parse_icici(month: dict) -> list[dict]:
    icici_dir = find_icici_dir(month)
    if icici_dir is None:
        print(f"  ICICI {month['date']}: no cached per-scheme files found "
              f"(expected under cache/mf_portfolios/icici_zip_{month['month_name'][:3]}/) -- SKIPPED")
        return []

    files = list(icici_dir.glob("*.xlsx"))
    all_rows = []
    matched_schemes = 0
    for target in ICICI_EQUITY_SCHEMES:
        candidates = [f for f in files if _fuzzy_match_scheme(f.stem, target)]
        if not candidates:
            print(f"  ICICI {month['date']}: no file matched for scheme '{target}'")
            continue
        f = candidates[0]
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ICICI {month['date']}: failed to open {f.name}: {e}")
            continue
        scheme_name = f"ICICI Prudential {target}"
        matched_schemes += 1
        for sheet_name in wb.sheetnames:
            if "deriv" in sheet_name.lower():
                continue  # separate derivatives sheet, different columns, holdings-lite
            rows = parse_holdings_sheet(wb[sheet_name], "ICICI", scheme_name, month["date"])
            all_rows.extend(rows)
            break  # first non-derivative sheet is the main holdings table
    print(f"  ICICI {month['date']}: {len(all_rows)} holding rows across {matched_schemes} equity schemes")
    return all_rows


def parse_hdfc(month: dict, include_elss: bool) -> list[dict]:
    hdfc_dir = CACHE_DIR / "hdfc"
    if not hdfc_dir.is_dir():
        print(f"  HDFC {month['date']}: no cached files found -- SKIPPED")
        return []

    port_dt = datetime.strptime(month["date"], "%Y-%m-%d")
    day_label = f"{port_dt.day} {month['month_title']} {port_dt.year}"
    schemes = list(HDFC_EQUITY_SCHEMES)
    if include_elss:
        schemes = schemes + [HDFC_ELSS_SCHEME]

    all_rows = []
    matched_schemes = 0
    for scheme in schemes:
        fname = f"Monthly HDFC {scheme} - {day_label}.xlsx"
        f = hdfc_dir / fname
        if not f.exists():
            print(f"  HDFC {month['date']}: missing file for scheme '{scheme}'")
            continue
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            print(f"  HDFC {month['date']}: failed to open {f.name}: {e}")
            continue
        scheme_name = f"HDFC {scheme}".replace("  ", " ").strip()
        matched_schemes += 1
        rows = parse_holdings_sheet(wb[wb.sheetnames[0]], "HDFC", scheme_name, month["date"])
        all_rows.extend(rows)
    print(f"  HDFC {month['date']}: {len(all_rows)} holding rows across {matched_schemes} equity schemes")
    return all_rows


# ===========================================================================
# SQLite + BigQuery
# ===========================================================================

def save_to_sqlite(df: pd.DataFrame, db_path: Path):
    conn = sqlite3.connect(db_path)
    try:
        existing_keys = set()
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='equity_portfolio_holdings'"
        )
        if cur.fetchone():
            existing = pd.read_sql(
                "SELECT DISTINCT amc, scheme_name, portfolio_date, isin FROM equity_portfolio_holdings",
                conn,
            )
            existing_keys = set(
                existing.itertuples(index=False, name=None)
            )
        key_cols = ["amc", "scheme_name", "portfolio_date", "isin"]
        if existing_keys:
            mask = ~df[key_cols].apply(tuple, axis=1).isin(existing_keys)
            df_new = df[mask]
        else:
            df_new = df
        df_new.to_sql("equity_portfolio_holdings", conn, if_exists="append", index=False)
        print(f"SQLite: inserted {len(df_new)} new rows (of {len(df)} parsed) into {db_path}")
        total = pd.read_sql("SELECT COUNT(*) as n FROM equity_portfolio_holdings", conn).iloc[0]["n"]
        print(f"SQLite: {db_path} now has {total} total rows")
    finally:
        conn.close()


def load_to_bigquery(df: pd.DataFrame, write_disposition: str = "WRITE_APPEND"):
    from google.cloud import bigquery

    client = bigquery.Client(project=PROJECT)
    client.create_dataset(f"{PROJECT}.{DATASET}", exists_ok=True)

    table_exists = True
    try:
        client.get_table(TABLE_ID)
    except Exception:
        table_exists = False

    if table_exists and write_disposition == "WRITE_APPEND":
        existing = client.query(
            f"SELECT DISTINCT amc, scheme_name, portfolio_date, isin FROM `{TABLE_ID}`"
        ).to_dataframe()
        if not existing.empty:
            key_cols = ["amc", "scheme_name", "portfolio_date", "isin"]
            df = df.merge(existing, on=key_cols, how="left", indicator=True)
            df = df[df["_merge"] == "left_only"].drop(columns="_merge")

    if df.empty:
        print("BigQuery: nothing new to load (all rows already present).")
        return

    job_config = bigquery.LoadJobConfig(write_disposition=write_disposition, autodetect=True)
    job = client.load_table_from_dataframe(df, TABLE_ID, job_config=job_config)
    job.result()
    print(f"BigQuery: loaded {len(df)} rows into {TABLE_ID}")


# ===========================================================================
# Main
# ===========================================================================

def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--months", type=int, default=3, help="how many of the most recent published months to process (default 3)")
    ap.add_argument("--out", default=str(DEFAULT_DB), help="SQLite database path")
    ap.add_argument("--to-bigquery", action="store_true", help=f"also load into {TABLE_ID}")
    ap.add_argument("--replace", action="store_true", help="WRITE_TRUNCATE instead of WRITE_APPEND for BigQuery")
    ap.add_argument("--no-download", action="store_true", help="skip network downloads, only use what's already cached")
    ap.add_argument("--no-elss", action="store_true", help="exclude HDFC ELSS Tax Saver Fund from the equity set")
    args = ap.parse_args()

    months = MONTHS[-args.months:]
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    all_rows = []
    for month in months:
        print(f"\n=== {month['date']} ===")

        if not args.no_download:
            sbi_path = download_sbi(month)
        else:
            sbi_path = CACHE_DIR / f"sbi_{month['date']}.xlsx"
            sbi_path = sbi_path if sbi_path.exists() else None
        all_rows.extend(parse_sbi(sbi_path, month["date"]))

        all_rows.extend(parse_icici(month))

        if not args.no_download:
            download_hdfc(month)
        all_rows.extend(parse_hdfc(month, include_elss=not args.no_elss))

    if not all_rows:
        print("\nNo rows parsed at all -- nothing to save.")
        return

    df = pd.DataFrame(all_rows)
    df = df.drop_duplicates(subset=["amc", "scheme_name", "portfolio_date", "isin", "instrument_name"])
    print(f"\nTotal parsed rows: {len(df)}")
    print(df.groupby(["amc", "portfolio_date"]).size())

    db_path = Path(args.out)
    save_to_sqlite(df, db_path)

    if args.to_bigquery:
        load_to_bigquery(df, write_disposition="WRITE_TRUNCATE" if args.replace else "WRITE_APPEND")


if __name__ == "__main__":
    main()
